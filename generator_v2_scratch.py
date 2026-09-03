#!/usr/bin/env python3
"""
MÉTODO 2: Construcción de Cero Optimizada - Libro Consolidado Completo
(generator_v2_scratch.py)

Lógica de extracción y construcción:
1. Plano estructural: Performance VP Retail 2026.xlsx se usa como esquema de estructura,
   estilos y fórmulas dinámicas locales (SUM, IF, etc.).
2. Alimentación desde fuentes externas: Extracción real desde archivos auxiliares de db/
   (Tarjetas, Vehicular, Convenios, Captaciones, Remesas, Fondos Mutuos, etc.).
3. Mapeo exhaustivo de enlaces: IDs [1]..[26] y nombres de hojas referenciadas con caché O(1).
4. Mapeo dinámico completo: Procesa TODAS las hojas del libro maestro.
5. Patrón Flyweight estricto: Asignación por referencia compartida, cero copy(),
   lectura en streaming con read_only=True para mantener RAM < 8 GB.
"""
import time
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class ScratchFlyweightFullGenerator:
    """Generador completo de Performance VP Retail desde cero con Flyweight"""

    DYNAMIC_FUNCTIONS = {
        'SUM', 'SUMIF', 'SUMIFS', 'AVERAGE', 'AVERAGEIF', 'AVERAGEIFS',
        'COUNT', 'COUNTA', 'COUNTIF', 'COUNTIFS', 'MAX', 'MIN', 'MEDIAN',
        'VLOOKUP', 'HLOOKUP', 'XLOOKUP', 'INDEX', 'MATCH', 'IF', 'IFS',
        'IFERROR', 'IFNA', 'AND', 'OR', 'NOT', 'DATE', 'DAY', 'MONTH',
        'YEAR', 'EOMONTH', 'CONCAT', 'CONCATENATE', 'TEXTJOIN'
    }

    # Mapeo exhaustivo de IDs de enlaces a archivos físicos en db/
    LINK_TO_FILE_MAP = {
        '2': 'Performance VP Retail 2024 - valores.xlsx',
        '3': 'Performance TC 2026.xlsm',
        '4': 'ROE_ROA_2026.xlsx',
        '5': 'ROE_ROA_2026.xlsx',
        '6': 'ROE_ROA_2026.xlsx',
        '7': 'ROE_ROA_2026.xlsx',
        '8': 'Performance Convenios 2026.xlsm',
        '9': 'Performance Vehicular 2026.xlsm',
        '10': 'Performance Préstamos_2026.xlsm',
        '11': 'Performance Adelanto_2026.xlsm',
        '12': 'Performance Hipotecario 2026.xlsm',
        '13': 'Performance Inmobiliaria 2026.xlsm',
        '14': 'Modelo VP Retail 2026.xlsx',
        '15': 'Performance Captaciones_2026.xlsm',
        '16': 'Performance Cuenta Sueldo_2026.xlsm',
        '17': 'Performance Remesas_2026.xlsm',
        '18': 'Performance Fondos Mutuos 2026.xlsx',
        '19': 'Performance VP Retail 2024 - valores.xlsx',
        '20': 'Modelo Consolidado VP Retail 2025.xlsx',
        '21': 'Performance VP Retail 2024 - valores.xlsx',
        '22': 'Performance VP Retail 2024 - valores.xlsx',
        '23': 'Performance VP Retail 2024 - valores.xlsx',
        '24': 'Performance VP Retail 2024 - valores.xlsx',
        '25': 'Performance VP Retail 2024 - valores.xlsx',
        '26': 'Performance VP Retail 2024 - valores.xlsx',
    }

    # Mapeo de respaldo por nombres de hoja a archivos fuente
    SHEET_TO_FILE_MAP = {
        'renta alta': 'Performance TC 2026.xlsm',
        'renta_alta': 'Performance TC 2026.xlsm',
        'masivo': 'Performance TC 2026.xlsm',
        'consumo inicial': 'Performance TC 2026.xlsm',
        'consumo_inicial': 'Performance TC 2026.xlsm',
        'tarjeta debito': 'Performance TC 2026.xlsm',
        'tarjeta_debito': 'Performance TC 2026.xlsm',
        'tarjeta empresarial': 'Performance TC 2026.xlsm',
        'desempeno finan tc cartera evo': 'Performance TC 2026.xlsm',
        'segmento': 'ROE_ROA_2026.xlsx',
        'segmento trimestral': 'ROE_ROA_2026.xlsx',
        'producto': 'ROE_ROA_2026.xlsx',
        'roe': 'ROE_ROA_2026.xlsx',
        'retail': 'ROE_ROA_2026.xlsx',
        'convenios': 'Performance Convenios 2026.xlsm',
        'desempeno financiero evol': 'Performance Convenios 2026.xlsm',
        'desempeno financiero (evol)': 'Performance Convenios 2026.xlsm',
        'vehicular': 'Performance Vehicular 2026.xlsm',
        'desempeno_financiero_evol': 'Performance Vehicular 2026.xlsm',
        'prestamos': 'Performance Préstamos_2026.xlsm',
        'adelanto': 'Performance Adelanto_2026.xlsm',
        'hipotecario': 'Performance Hipotecario 2026.xlsm',
        'desempeno_financiero (evol)-seg': 'Performance Hipotecario 2026.xlsm',
        'inmobiliaria': 'Performance Inmobiliaria 2026.xlsm',
        'resultados': 'Performance Inmobiliaria 2026.xlsm',
        'captaciones': 'Performance Captaciones_2026.xlsm',
        'cuenta sueldo': 'Performance Cuenta Sueldo_2026.xlsm',
        'remesas': 'Performance Remesas_2026.xlsm',
        'resumen': 'Performance Fondos Mutuos 2026.xlsx',
        'fondos mutuos': 'Performance Fondos Mutuos 2026.xlsx',
        'producto-evolutivo': 'Performance VP Retail 2024 - valores.xlsx',
        'evolutivo': 'Performance Adelanto_2026.xlsm',
    }

    # Archivos alternativos de respaldo si el principal .xlsm no estuviera
    FALLBACK_FILES = {
        'Performance TC 2026.xlsm': 'Performance Tarjetas.xlsx',
        'Performance Convenios 2026.xlsm': 'Performance Convenios.xlsx',
        'Performance Préstamos_2026.xlsm': 'Performance Préstamos.xlsx',
        'Performance Hipotecario 2026.xlsm': 'Performance Hipotecario.xlsx',
        'Performance Captaciones_2026.xlsm': 'Performance Captaciones.xlsx',
        'Performance Cuenta Sueldo_2026.xlsm': 'Performance Cta Sueldo.xlsx',
    }

    def __init__(self, db_dir=None):
        self.db_dir = Path(db_dir) if db_dir else Path(__file__).parent / "db"
        self.reference_path = self.db_dir / "Performance VP Retail 2026.xlsx"
        self.sources_cache = {}  # (file_name, sheet_clean) -> {coord: value}
        self.loaded_workbooks = {}  # file_path -> openpyxl.Workbook (read_only)
        self.sheet_metadata = {}  # sheet_name -> {cols: [...], merges: [...]}
        self.stats = {
            'sheets_generated': 0,
            'dynamic_formulas': 0,
            'resolved_external_values': 0,
            'cached_values': 0,
            'formats_applied': 0,
            'cells_written': 0,
        }

    def _normalize_name(self, name):
        """Normaliza nombres de hoja para búsquedas insensibles a caracteres especiales"""
        import unicodedata
        nfkd = unicodedata.normalize('NFKD', name.strip("'\" "))
        clean = "".join([c for c in nfkd if not unicodedata.combining(c)])
        return clean.replace('_', ' ').replace('-', ' ').lower()

    def _preload_sheet_metadata(self):
        """Extrae de forma instantánea dimensiones y celdas combinadas vía XML"""
        if not self.reference_path.exists():
            return
        try:
            with zipfile.ZipFile(self.reference_path, 'r') as zin:
                # Leer mapeo de hojas en workbook.xml
                wb_xml = zin.read('xl/workbook.xml')
                root = ET.fromstring(wb_xml)
                ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                
                # Leer relaciones de hojas
                rels_xml = zin.read('xl/_rels/workbook.xml.rels')
                rels_root = ET.fromstring(rels_xml)
                rel_map = {r.attrib['Id']: r.attrib['Target'] for r in rels_root}

                for s in root.findall('main:sheets/main:sheet', ns):
                    sname = s.attrib['name']
                    r_id = s.attrib['{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id']
                    target_file = 'xl/' + rel_map[r_id]
                    
                    if target_file in zin.namelist():
                        s_data = zin.read(target_file)
                        s_root = ET.fromstring(s_data)
                        
                        cols = []
                        for col in s_root.findall('main:cols/main:col', ns):
                            min_c = int(col.attrib.get('min', 1))
                            max_c = int(col.attrib.get('max', 1))
                            w = float(col.attrib.get('width', 10))
                            cols.append((min_c, max_c, w))

                        merges = []
                        for m in s_root.findall('main:mergeCells/main:mergeCell', ns):
                            merges.append(m.attrib.get('ref'))

                        self.sheet_metadata[sname] = {
                            'cols': cols,
                            'merges': merges
                        }
        except Exception:
            pass

    def _get_source_sheet_data(self, link_id, raw_sheet_name):
        """Carga y almacena en caché O(1) una hoja fuente completa"""
        sheet_clean = self._normalize_name(raw_sheet_name)
        
        filename = None
        if link_id and link_id in self.LINK_TO_FILE_MAP:
            filename = self.LINK_TO_FILE_MAP[link_id]
        if not filename or not (self.db_dir / filename).exists():
            filename = self.SHEET_TO_FILE_MAP.get(sheet_clean)

        if not filename:
            return None

        file_path = self.db_dir / filename
        if not file_path.exists() and filename in self.FALLBACK_FILES:
            fallback = self.FALLBACK_FILES[filename]
            if (self.db_dir / fallback).exists():
                file_path = self.db_dir / fallback
                filename = fallback

        if not file_path.exists():
            return None

        cache_key = (filename.lower(), sheet_clean)
        if cache_key in self.sources_cache:
            return self.sources_cache[cache_key]

        try:
            if file_path not in self.loaded_workbooks:
                self.loaded_workbooks[file_path] = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            wb_src = self.loaded_workbooks[file_path]

            target_ws = None
            for sn in wb_src.sheetnames:
                norm_sn = self._normalize_name(sn)
                if norm_sn == sheet_clean or norm_sn.replace(' ', '') == sheet_clean.replace(' ', ''):
                    target_ws = wb_src[sn]
                    break

            if target_ws is None:
                self.sources_cache[cache_key] = None
                return None

            sheet_dict = {}
            for r_idx, row in enumerate(target_ws.iter_rows(values_only=True), 1):
                for c_idx, val in enumerate(row, 1):
                    if val is not None:
                        coord = f"{get_column_letter(c_idx)}{r_idx}"
                        sheet_dict[coord] = val

            self.sources_cache[cache_key] = sheet_dict
            return sheet_dict
        except Exception:
            self.sources_cache[cache_key] = None
            return None

    def is_dynamic_formula(self, formula):
        """Determina si la fórmula debe mantenerse dinámica"""
        if not isinstance(formula, str) or not formula.startswith('='):
            return False
        if '[' in formula and ']' in formula:
            return False
        formula_upper = formula.upper()
        if any(f"{fn}(" in formula_upper for fn in self.DYNAMIC_FUNCTIONS):
            return True
        if any(op in formula for op in ['+', '-', '*', '/', '^']):
            return True
        return False

    def resolve_cell_value(self, formula_val, cached_val):
        """Resuelve el valor exacto según reglas de paridad estricta"""
        # 1. Fórmula dinámica local
        if self.is_dynamic_formula(formula_val):
            self.stats['dynamic_formulas'] += 1
            return formula_val

        # 2. Fórmula con referencia externa ([N]Sheet!Cell)
        if isinstance(formula_val, str) and formula_val.startswith('=') and '[' in formula_val:
            m = re.search(r'\[(\d+)\]([^!]+)!([A-Z]+\d+)', formula_val)
            if m:
                link_id, sheet_name, cell_addr = m.groups()
                sheet_data = self._get_source_sheet_data(link_id, sheet_name)
                if sheet_data and cell_addr in sheet_data:
                    self.stats['resolved_external_values'] += 1
                    return sheet_data[cell_addr]

            # Fallback a valor en caché
            if cached_val is not None:
                self.stats['cached_values'] += 1
                return cached_val
            return None

        # 3. Valor simple o constante
        if cached_val is not None:
            self.stats['cached_values'] += 1
            return cached_val

        return None

    def generate(self, output_path):
        """Construye el libro completo consolidado desde cero"""
        total_start = time.time()
        output_path = Path(output_path)

        print("\n" + "="*70)
        print("🚀 GENERADOR V2 COMPLETO: CONSTRUCCIÓN DESDE CERO (FLYWEIGHT)")
        print("="*70)
        print("  ✅ Estructura y fórmulas dinámicas: Performance VP Retail 2026.xlsx")
        print("  ✅ Datos reales extraídos: Archivos auxiliares de db/")
        print("  ✅ Patrón Flyweight para estilos (sin copy)")
        print("  ✅ Lectura en streaming (read_only=True)\n")

        # Precargar metadata de dimensiones y merges
        self._preload_sheet_metadata()

        wb_output = Workbook()
        wb_output.remove(wb_output.active)  # Eliminar hoja inicial vacía

        # Abrir libro de referencia en streaming
        wb_formulas = openpyxl.load_workbook(self.reference_path, data_only=False, read_only=True)
        wb_values = openpyxl.load_workbook(self.reference_path, data_only=True, read_only=True)

        sheet_times = {}
        all_sheets = wb_formulas.sheetnames
        print(f"Iniciando procesamiento de {len(all_sheets)} hojas...\n")

        for idx, sheet_name in enumerate(all_sheets, 1):
            t_sheet_start = time.time()
            ws_output = wb_output.create_sheet(title=sheet_name)
            ws_formulas = wb_formulas[sheet_name]
            ws_values = wb_values[sheet_name]

            formulas_iter = ws_formulas.iter_rows(values_only=False)
            values_iter = ws_values.iter_rows(values_only=True)

            row_idx = 1
            written_in_sheet = 0
            sheet_dyn_formulas = 0
            sheet_ext_resolved = 0

            for f_row, v_row in zip(formulas_iter, values_iter):
                for col_idx, (src_cell, v_val) in enumerate(zip(f_row, v_row), 1):
                    f_val = src_cell.value
                    has_style = bool(src_cell.font or src_cell.fill or src_cell.border or src_cell.alignment or src_cell.number_format != 'General')

                    if f_val is not None or v_val is not None or has_style:
                        final_val = self.resolve_cell_value(f_val, v_val)
                        target_cell = ws_output.cell(row=row_idx, column=col_idx)

                        if final_val is not None:
                            target_cell.value = final_val

                        # Patrón Flyweight: Asignación por referencia directa sin copy()
                        if src_cell.font:
                            target_cell.font = src_cell.font
                        if src_cell.fill:
                            target_cell.fill = src_cell.fill
                        if src_cell.border:
                            target_cell.border = src_cell.border
                        if src_cell.alignment:
                            target_cell.alignment = src_cell.alignment
                        if src_cell.number_format:
                            target_cell.number_format = src_cell.number_format

                        self.stats['formats_applied'] += 1
                        written_in_sheet += 1
                row_idx += 1

            # Aplicar anchos de columna y celdas combinadas si existen en metadata
            if sheet_name in self.sheet_metadata:
                meta = self.sheet_metadata[sheet_name]
                for min_c, max_c, w in meta.get('cols', []):
                    for c in range(min_c, max_c + 1):
                        ws_output.column_dimensions[get_column_letter(c)].width = w
                for m_ref in meta.get('merges', []):
                    try:
                        ws_output.merge_cells(m_ref)
                    except Exception:
                        pass

            self.stats['cells_written'] += written_in_sheet
            self.stats['sheets_generated'] += 1
            elapsed_sheet = time.time() - t_sheet_start
            sheet_times[sheet_name] = elapsed_sheet
            
            print(f"  [{idx:02d}/{len(all_sheets):02d}] {sheet_name:<26} │ {row_idx-1:>5} filas │ {written_in_sheet:>7,} celdas │ {elapsed_sheet:>6.2f} s")

        wb_formulas.close()
        wb_values.close()
        for wb in self.loaded_workbooks.values():
            try:
                wb.close()
            except Exception:
                pass

        print("\nGuardando libro consolidado...")
        t_save = time.time()
        wb_output.save(output_path)
        wb_output.close()
        save_time = time.time() - t_save
        total_time = time.time() - total_start

        print("\n" + "="*70)
        print("  RESUMEN FINAL: PERFORMANCE VP RETAIL COMPLETO")
        print("="*70)
        print(f"  Hojas procesadas:          {self.stats['sheets_generated']} de {len(all_sheets)}")
        print(f"  Total celdas construidas:  {self.stats['cells_written']:,}")
        print(f"  Fórmulas dinámicas:        {self.stats['dynamic_formulas']:,}")
        print(f"  Enlaces externos resueltos:{self.stats['resolved_external_values']:,}")
        print(f"  Formatos Flyweight:        {self.stats['formats_applied']:,}")
        print(f"  Guardado final en disco:   {save_time:.2f} s")
        print("-"*70)
        print(f"  ⏱️  TIEMPO TOTAL:            {total_time:.2f} s ({total_time/60:.2f} min)")
        print(f"  💾 Archivo generado:       {output_path.name} ({output_path.stat().st_size / (1024*1024):.2f} MB)")
        print("="*70 + "\n")

        return {
            'success': True,
            'total_time': total_time,
            'sheet_times': sheet_times,
            'file_size_mb': output_path.stat().st_size / (1024*1024),
            'stats': self.stats
        }

if __name__ == "__main__":
    db_path = Path(__file__).parent / "db"
    out_file = db_path / "Performance_VP_Retail_2026_COMPLETO_V2.xlsx"
    generator = ScratchFlyweightFullGenerator(db_path)
    generator.generate(out_file)
