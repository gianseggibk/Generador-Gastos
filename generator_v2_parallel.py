#!/usr/bin/env python3
"""
GENERADOR V2 PARALELO MULTI-HILO (generator_v2_parallel.py)
Máxima velocidad aprovechando todos los núcleos/hilos del CPU.
100% compatible con Windows y macOS sin restricciones de IPC.
"""
import time
import os
import re
import sys
import shutil
import zipfile
import tempfile
import unicodedata
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.utils import get_column_letter

LINK_TO_FILE_MAP = {
    '2': 'Performance VP Retail 2024 - valores.xlsx',
    '3': 'Performance TC 2026.xlsm',
    '4': 'ROE_ROA_2026.xlsx',
    '5': 'ROE_ROA_2026.xlsx',
    '6': 'ROE_ROA_2026.xlsx',
    '7': 'ROE_ROA_2026.xlsx',
    '8': 'Performance Convenios 2026.xlsm',
    '9': 'Performance Vehicular 2026.xlsm',
    '10': 'Performance Préstamos_2026.xlsm',
    '11': 'Performance Adelanto_2026.xlsm',
    '12': 'Performance Hipotecario 2026.xlsm',
    '13': 'Performance Inmobiliaria 2026.xlsm',
    '14': 'Modelo VP Retail 2026.xlsx',
    '15': 'Performance Captaciones_2026.xlsm',
    '16': 'Performance Cuenta Sueldo_2026.xlsm',
    '17': 'Performance Remesas_2026.xlsm',
    '18': 'Performance Fondos Mutuos 2026.xlsx',
    '19': 'Performance VP Retail 2024 - valores.xlsx',
    '20': 'Modelo Consolidado VP Retail 2025.xlsx',
    '21': 'Performance VP Retail 2024 - valores.xlsx',
    '22': 'Performance VP Retail 2024 - valores.xlsx',
    '23': 'Performance VP Retail 2024 - valores.xlsx',
    '24': 'Performance VP Retail 2024 - valores.xlsx',
    '25': 'Performance VP Retail 2024 - valores.xlsx',
    '26': 'Performance VP Retail 2024 - valores.xlsx',
}

SHEET_TO_FILE_MAP = {
    'renta alta': 'Performance TC 2026.xlsm',
    'renta_alta': 'Performance TC 2026.xlsm',
    'masivo': 'Performance TC 2026.xlsm',
    'consumo inicial': 'Performance TC 2026.xlsm',
    'consumo_inicial': 'Performance TC 2026.xlsm',
    'tarjeta debito': 'Performance TC 2026.xlsm',
    'tarjeta_debito': 'Performance TC 2026.xlsm',
    'tarjeta empresarial': 'Performance TC 2026.xlsm',
    'desempeno finan tc cartera evo': 'Performance TC 2026.xlsm',
    'segmento': 'ROE_ROA_2026.xlsx',
    'segmento trimestral': 'ROE_ROA_2026.xlsx',
    'producto': 'ROE_ROA_2026.xlsx',
    'roe': 'ROE_ROA_2026.xlsx',
    'retail': 'ROE_ROA_2026.xlsx',
    'convenios': 'Performance Convenios 2026.xlsm',
    'desempeno financiero evol': 'Performance Convenios 2026.xlsm',
    'desempeno financiero (evol)': 'Performance Convenios 2026.xlsm',
    'vehicular': 'Performance Vehicular 2026.xlsm',
    'desempeno_financiero_evol': 'Performance Vehicular 2026.xlsm',
    'prestamos': 'Performance Préstamos_2026.xlsm',
    'adelanto': 'Performance Adelanto_2026.xlsm',
    'hipotecario': 'Performance Hipotecario 2026.xlsm',
    'desempeno_financiero (evol)-seg': 'Performance Hipotecario 2026.xlsm',
    'inmobiliaria': 'Performance Inmobiliaria 2026.xlsm',
    'resultados': 'Performance Inmobiliaria 2026.xlsm',
    'captaciones': 'Performance Captaciones_2026.xlsm',
    'cuenta sueldo': 'Performance Cuenta Sueldo_2026.xlsm',
    'remesas': 'Performance Remesas_2026.xlsm',
    'resumen': 'Performance Fondos Mutuos 2026.xlsx',
    'fondos mutuos': 'Performance Fondos Mutuos 2026.xlsx',
    'producto-evolutivo': 'Performance VP Retail 2024 - valores.xlsx',
    'evolutivo': 'Performance Adelanto_2026.xlsm',
}

FALLBACK_FILES = {
    'Performance TC 2026.xlsm': 'Performance Tarjetas.xlsx',
    'Performance Convenios 2026.xlsm': 'Performance Convenios.xlsx',
    'Performance Préstamos_2026.xlsm': 'Performance Préstamos.xlsx',
    'Performance Hipotecario 2026.xlsm': 'Performance Hipotecario.xlsx',
    'Performance Captaciones_2026.xlsm': 'Performance Captaciones.xlsx',
    'Performance Cuenta Sueldo_2026.xlsm': 'Performance Cta Sueldo.xlsx',
}

def normalize_name(name):
    nfkd = unicodedata.normalize('NFKD', str(name).strip("'\" "))
    clean = "".join([c for c in nfkd if not unicodedata.combining(c)])
    return clean.replace('_', ' ').replace('-', ' ').lower()

class ThreadParallelConsolidator:
    """Consolidador paralelo multi-hilo con caché de memoria compartida"""

    def __init__(self, db_dir=None):
        self.db_dir = Path(db_dir) if db_dir else Path(__file__).parent / "db"
        self.reference_path = self.db_dir / "Performance VP Retail 2026.xlsx"
        self.sources_cache = {}
        self.loaded_wbs = {}

    def get_source_val(self, link_id, sheet_name, cell_addr):
        sheet_clean = normalize_name(sheet_name)
        fname = LINK_TO_FILE_MAP.get(str(link_id))
        if not fname or not (self.db_dir / fname).exists():
            fname = SHEET_TO_FILE_MAP.get(sheet_clean)

        if not fname:
            return None

        fpath = self.db_dir / fname
        if not fpath.exists() and fname in FALLBACK_FILES:
            fallback = FALLBACK_FILES[fname]
            if (self.db_dir / fallback).exists():
                fpath = self.db_dir / fallback
                fname = fallback

        if not fpath.exists():
            return None

        cache_key = (fname.lower(), sheet_clean)
        if cache_key not in self.sources_cache:
            try:
                if fpath not in self.loaded_wbs:
                    self.loaded_wbs[fpath] = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                wb = self.loaded_wbs[fpath]

                target_ws = None
                for sn in wb.sheetnames:
                    norm_sn = normalize_name(sn)
                    if norm_sn == sheet_clean or norm_sn.replace(' ', '') == sheet_clean.replace(' ', ''):
                        target_ws = wb[sn]
                        break

                if target_ws is None:
                    self.sources_cache[cache_key] = None
                else:
                    s_dict = {}
                    for r_idx, row in enumerate(target_ws.iter_rows(values_only=True), 1):
                        for c_idx, val in enumerate(row, 1):
                            if val is not None:
                                s_dict[f"{get_column_letter(c_idx)}{r_idx}"] = val
                    self.sources_cache[cache_key] = s_dict
            except Exception:
                self.sources_cache[cache_key] = None

        sheet_dict = self.sources_cache.get(cache_key)
        if sheet_dict and cell_addr in sheet_dict:
            return sheet_dict[cell_addr]
        return None

    def process_sheet_content(self, sname, xml_text):
        """Procesa el contenido XML de una hoja reemplazando enlaces externos"""
        t0 = time.time()
        dyn_count = 0
        ext_count = 0
        cells_count = 0

        cell_regex = re.compile(r'<c\s+([^>]*?)>(.*?)</c>|<c\s+([^>]*?)/>', re.DOTALL)

        def cell_replacer(m):
            nonlocal dyn_count, ext_count, cells_count
            cells_count += 1
            if m.group(3):
                return m.group(0)

            attrs = m.group(1)
            body = m.group(2)

            f_match = re.search(r'<f([^>]*)>(.*?)</f>', body, re.DOTALL)
            if f_match:
                formula = f_match.group(2)

                if '[' in formula and ']' in formula:
                    ext_m = re.search(r'\[(\d+)\]([^!]+)!([A-Z]+\d+)', formula)
                    if ext_m:
                        lid, s_ref, c_addr = ext_m.groups()
                        val = self.get_source_val(lid, s_ref, c_addr)
                        if val is not None:
                            ext_count += 1
                            clean_attrs = re.sub(r'\s+t="[^"]*"', '', attrs)
                            if isinstance(val, (int, float)):
                                return f'<c {clean_attrs}><v>{val}</v></c>'
                            else:
                                return f'<c {clean_attrs} t="inlineStr"><is><t>{val}</t></is></c>'

                    v_match = re.search(r'<v>(.*?)</v>', body)
                    clean_attrs = re.sub(r'\s+t="[^"]*"', '', attrs)
                    if v_match:
                        return f'<c {clean_attrs}><v>{v_match.group(1)}</v></c>'
                    return f'<c {attrs}/>'

                dyn_count += 1
                return m.group(0)

            return m.group(0)

        new_text = cell_regex.sub(cell_replacer, xml_text)
        elapsed = time.time() - t0

        return {
            'success': True,
            'sheet_name': sname,
            'processed_xml': new_text.encode('utf-8'),
            'dynamic_formulas': dyn_count,
            'resolved_external': ext_count,
            'cells_processed': cells_count,
            'elapsed': elapsed
        }

    def generate(self, output_path, max_workers=None):
        total_start = time.time()
        output_path = Path(output_path)
        
        num_threads = max_workers or min(os.cpu_count() or 4, 8)

        print("\n" + "="*72)
        print("⚡ GENERADOR V2 PARALELO MULTI-HILO (PERFORMANCE VP RETAIL)")
        print("="*72)
        print(f"  Hilos concurrentes asignados:   {num_threads}")
        print(f"  Archivo maestro de estructura:  {self.reference_path.name}")
        print(f"  Fuentes de datos auxiliares:    Carpeta db/ (*.xlsm, *.xlsx)\n")

        # 1. Leer archivo base en memoria
        t_read = time.time()
        with zipfile.ZipFile(self.reference_path, 'r') as zin:
            wb_xml = zin.read('xl/workbook.xml').decode('utf-8')
            root = ET.fromstring(wb_xml)
            ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            
            rels_xml = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')
            rels_root = ET.fromstring(rels_xml)
            rel_map = {r.attrib['Id']: r.attrib['Target'] for r in rels_root}

            sheet_entries = []
            for s in root.findall('main:sheets/main:sheet', ns):
                sname = s.attrib['name']
                r_id = s.attrib['{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id']
                rel_target = rel_map[r_id]
                sheet_entries.append((sname, 'xl/' + rel_target))

            sheet_texts = {}
            for sname, target_file in sheet_entries:
                sheet_texts[sname] = zin.read(target_file).decode('utf-8')

            non_sheet_files = {}
            sheet_target_names = {t for _, t in sheet_entries}
            for item in zin.infolist():
                if item.filename not in sheet_target_names:
                    non_sheet_files[item.filename] = zin.read(item.filename)

        print(f"Estructura leída en {time.time()-t_read:.2f}s.")
        print(f"Procesando {len(sheet_entries)} hojas en paralelo sobre {num_threads} hilos...\n")

        # 2. Ejecutar procesamiento paralelo con ThreadPoolExecutor
        t_pool = time.time()
        with ThreadPoolExecutor(max_workers=num_threads) as executor:
            futures = [
                executor.submit(self.process_sheet_content, sname, sheet_texts[sname])
                for sname, _ in sheet_entries
            ]
            results = [f.result() for f in futures]
        pool_time = time.time() - t_pool

        # 3. Cerrar libros abiertos
        for wb in self.loaded_wbs.values():
            try:
                wb.close()
            except Exception:
                pass

        total_dyn = 0
        total_ext = 0
        total_cells = 0
        processed_sheets = {}

        for idx, (res, (sname, target_file)) in enumerate(zip(results, sheet_entries), 1):
            processed_sheets[target_file] = res['processed_xml']
            total_dyn += res['dynamic_formulas']
            total_ext += res['resolved_external']
            total_cells += res['cells_processed']
            print(f"  [{idx:02d}/{len(sheet_entries):02d}] {sname:<26} │ {res['dynamic_formulas']:>7,} dinám. │ {res['resolved_external']:>6,} ext. │ {res['elapsed']:>5.2f} s")

        # 4. Re-empaquetar ZIP comprimido
        print(f"\nEnsamblando archivo consolidado .xlsx...")
        t_zip = time.time()
        with zipfile.ZipFile(output_path, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for fname, data in non_sheet_files.items():
                zout.writestr(fname, data)
            for fname, data in processed_sheets.items():
                zout.writestr(fname, data)

        zip_time = time.time() - t_zip
        total_time = time.time() - total_start

        print("\n" + "="*72)
        print("  RESUMEN FINAL: CONSOLIDADO COMPLETO (PARALELO)")
        print("="*72)
        print(f"  Hojas procesadas:          {len(sheet_entries)} de {len(sheet_entries)} (100% completas)")
        print(f"  Total celdas construidas:  {total_cells:,}")
        print(f"  Fórmulas dinámicas:        {total_dyn:,}")
        print(f"  Enlaces resueltos O(1):    {total_ext:,}")
        print(f"  Tiempo procesamiento:      {pool_time:.2f} s ({num_threads} hilos concurrentes)")
        print(f"  Ensamblado ZIP final:      {zip_time:.2f} s")
        print("-"*70)
        print(f"  ⏱️  TIEMPO TOTAL:            {total_time:.2f} s ({total_time/60:.2f} min)")
        print(f"  💾 Archivo generado:       {output_path.name} ({output_path.stat().st_size / (1024*1024):.2f} MB)")
        print("="*72 + "\n")

        return {
            'success': True,
            'total_time': total_time,
            'file_size_mb': output_path.stat().st_size / (1024*1024),
            'stats': {
                'total_sheets': len(sheet_entries),
                'dynamic_formulas': total_dyn,
                'resolved_external': total_ext,
                'cells_processed': total_cells
            }
        }

if __name__ == "__main__":
    db_path = Path(__file__).parent / "db"
    out_file = db_path / "Performance_VP_Retail_2026_PARALELO.xlsx"
    generator = ThreadParallelConsolidator(db_path)
    generator.generate(out_file)
